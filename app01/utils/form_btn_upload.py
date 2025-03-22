"""
自定义的模态框--从excel导入数据，使用方法：
# 在视图函数中：
    def salesdata_addform(request):
        from openpyxl import load_workbook
        # 1.获取用户上传的文件对象
        file_obj = request.FILES.get("upload_file")
        if not file_obj:
            return JsonResponse({"status": False, "error": "请上传文件"})

        # 2.对象传递给openpyxl，又openpyxl读取文件的内容
        try:
            wb = load_workbook(file_obj)
            sheet = wb.active  # 取第一个工作表
        except Exception:
            return JsonResponse({"status": False, "error": "Excel 解析失败"})

        # 3.循环获取每一行数据
        sales_data_list = []
        error_list = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            date_dict = {
                'date': row[0].value,
                'order_number': row[1].value,
            }
            # 用 ModelForm 进行校验
            try:
                form = SalesDataModelForm(date_dict)  # 用 ModelForm 进行校验
                if form.is_valid():
                    sales_data_list.append(models.SalesData(**date_dict))
                else:
                    error_list.append(f"第 {idx} 行错误: {form.errors}")
            except Exception as error:
                error_list.append(f"第 {idx} 行错误: <ul><li>{error}(查看日期格式是否规范化，标准2025-01-20)</li></ul>")

        # 如果有错误，返回错误信息
        if error_list:
            return JsonResponse({"status": False, "error": error_list})

        try:
            models.SalesData.objects.bulk_create(sales_data_list)
        except Exception as error:
            print(error)
            return JsonResponse({"status": False, "error": f"数据导入数据库失败: {str(error)}"})
        return JsonResponse({"status": True})

# 在HTML页面中：（依赖jquery）
    # 1、模态框弹窗激活按钮
        <a class="btn btn-primary" id="btnUploadBar"><span class="glyphicon glyphicon-paste" aria-hidden="true"></span>导入数据</a>
    # 2、模态框弹窗组件，放在html_body随意一个位置
    {{ upload_Modal }}
    # 3、导入模态框的js
    {% block js %}
        {{ upload_js }}
    {% endblock %}
"""

from django.middleware.csrf import get_token
from django.utils.safestring import mark_safe


class FormBtnUpload:

    def __init__(self, request, upload_post_url, modal_tittle, modal_label):
        """
        属性：
        :param request: 请求的对象
        :param upload_post_url: 提交表单的url
        :param modal_tittle: 模态框标题
        :param modal_label: 上传文件按钮上方的提示内容
        """
        self.request = request
        self.upload_post_url = upload_post_url
        self.modal_tittle = modal_tittle
        self.modal_label = modal_label

    def html_modal(self):
        """
        批量上传 模态框
        """
        csrf_token = get_token(self.request)  # 获取 CSRF 令牌
        page = f"""
            <div class="modal fade" id="myModal_uploadBar" tabindex="-1" role="dialog">
                <div class="modal-dialog modal-lg" role="document">
                    <div class="modal-content">
                        <div class="modal-header">
                            <button type="button" class="close" data-dismiss="modal" aria-label="Close"><span aria-hidden="true">×</span></button>
                            <h4 class="modal-title">{self.modal_tittle}</h4>
                        </div>
                        <form id="formUploadFile" enctype="multipart/form-data" novalidate>
                            <div class="modal-body">
                                <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                                <div class="form-group">
                                    <div class="alert alert-danger hide" role="alert" id="myModal_uploadBar_error" style="margin-top: 10px"></div>
                                    <label style="margin-bottom: 10px">{self.modal_label}</label>
                                    <input type="file" name="upload_file" >
                                </div>
                            </div>
                            <div class="modal-footer">
                                <button type="button" class="btn btn-default" data-dismiss="modal">取 消</button>
                                <button id="btnUpload" class="btn btn-primary">上 传</button>
                            </div>
                        </form>
                    </div><!-- /.modal-content -->
                </div><!-- /.modal-dialog -->
            </div><!-- /.modal -->
        """
        page_string = mark_safe(page)  # 导入django的mark_safe模块，字符串才会写进html页面中
        return page_string

    def js(self):
        page_str_list = []
        page = f"""
            <script type="text/javascript">
                $(function () {{
                    bindBtnUploadBarEvent();
                    bindBtnUploadSaveEvent();
                }})
        
                // 弹出对话框
                function bindBtnUploadBarEvent() {{
                    $("#btnUploadBar").click(function () {{
                        // 点击该按钮，显示对话框
                        $("#myModal_uploadBar").modal("show");
                        $("#myModal_uploadBar_error").addClass("hide");
                    }});
                }}
        
                // 批量上传的对话框的保存按钮
                function bindBtnUploadSaveEvent() {{
                    $("#btnUpload").click(function () {{
                        event.preventDefault();  // 阻止表单的默认提交行为

                        var originalText = $(this).text();  // 保存原始按钮文字
                        var originalClass = $(this).attr("class");  // 保存原始 class
                        $(this).text('导入中...');  // 更改按钮文字
                        $(this).removeClass("btn-primary").addClass("btn-warning");  // 修改按钮样式为“警告”按钮
                        
                        var formData = new FormData($("#formUploadFile")[0]); // 获取表单数据
                        $.ajax({{
                            url: "{self.upload_post_url}",
                            type: "post",
                            data: formData,
                            processData: false,  // 不处理数据，让 FormData 以二进制形式传输
                            contentType: false,  // 让浏览器自动设置 Content-Type
                            success: function (res) {{  // {{"status": False, "error": form.errors}}
                                console.log(res);
                                if (res.status) {{
                                    alert("导入成功");
                                    location.reload();
                                }} else {{
                                    var $field = $("#myModal_uploadBar").find("#myModal_uploadBar_error");  // 确保只选当前模态框里的字段
                                    $field.html(res.error);
                                    $("#myModal_uploadBar_error").removeClass("hide");
                                }}
                            }},
                            complete: function () {{
                                $("#btnUpload").text(originalText);  // 恢复原始按钮文字
                                $("#btnUpload").attr("class", originalClass);  // 恢复原始 class
                            }}
                        }});
                    }});
                }}
            </script>
        """
        page_str_list.append(page)
        page_string = mark_safe("".join(page_str_list))  # 导入django的mark_safe模块，字符串才会写进html页面中
        return page_string
