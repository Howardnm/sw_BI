from django.http import JsonResponse
from django.shortcuts import render, HttpResponse, redirect

from app01 import models
from app01.utils.encrypt import md5
from app01.utils.form import SalesIndicatorModelForm
from app01.utils.form_btn_add import FormBtnAdd
from app01.utils.form_btn_delete import FormBtnDelete
from app01.utils.form_btn_delete_all import FormBtnDelAll
from app01.utils.form_btn_edit import FormBtnEdit
from app01.utils.form_btn_upload import FormBtnUpload
from app01.utils.pagination import Pagination
from app01.utils.form_btn_search import FormBtnSearch

url_path = "salesindicator"


def list(request):
    """ 销售大信息列表 """
    form = SalesIndicatorModelForm()  # 引入表模板
    search_bar = FormBtnSearch(request, form)  # 搜索框组件对象
    queryset = models.SalesIndicator.objects.filter(**search_bar.filter()).order_by("-id")
    page_obj = Pagination(request, queryset, "page", 15)  # 翻页组件对象
    add_obj = FormBtnAdd(request, form, f"/{url_path}/add", "新增数据")  # 添加框组件对象
    edit_obj = FormBtnEdit(request, form, f"/{url_path}/edit_detail", f"/{url_path}/edit", "编辑数据")  # 编辑框组件对象
    del_obj = FormBtnDelete(f"/{url_path}/delete")  # 删除框组件对象
    del_all_obj = FormBtnDelAll(request, f"/{url_path}/delete_all", "请输入管理员密码")  # 一键删除框组件对象

    upload_obj = FormBtnUpload(request, f"/{url_path}/addform", "/static/files/salesindicator_addform.xlsx", "导入销售信息", "选择规范化的excel文件进行导入")  # excel批量上传框组件对象
    context = {
        # 搜索框组件
        "search_Modal": search_bar.html_modal(),
        "search_js": search_bar.js(),
        # 新增数据框组件
        "add_Modal": add_obj.html_modal(),
        "add_js": add_obj.js(),
        # 编辑数据框组件
        "edit_Modal": edit_obj.html_modal(),
        "edit_js": edit_obj.js(),
        # 删除框组件
        "delete_Modal": del_obj.html_modal(),
        "delete_js": del_obj.js(),
        # 一键删除框组件
        "delete_all_Modal": del_all_obj.html_modal(),
        "delete_all_js": del_all_obj.js(),
        # excel批量上传框组件
        "upload_Modal": upload_obj.html_modal(),
        "upload_js": upload_obj.js(),
        # 翻页组件
        "queryset": page_obj.page_queryset,  # 分完页的数据
        "page_string": page_obj.html(),  # html页码
    }
    return render(request, f"{url_path}_list.html", context)


def add(request):
    """ 添加个人业绩信息 """
    form = SalesIndicatorModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})  # 直接传入dict，django会自动转json
    return JsonResponse({"status": False, "error": form.errors})


def edit(request):
    """ 编辑信息 """
    uid = request.GET.get("uid")
    obj = models.SalesIndicator.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({"status": False, "tips": "无法编辑，该行数据不存在，请刷新重试"})
    form = SalesIndicatorModelForm(data=request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


def edit_detail(request):
    """ 获取要编辑的数据 """
    uid = request.GET.get("uid")
    obj_dict = models.SalesIndicator.objects.filter(id=uid).values().first()  # 获得的是一个字典
    obj_dict["name"] = obj_dict["name_id"]  # 由于name是关联字段，sql中只有name_id字段，但是form是name字段，所以要补充一个
    if not obj_dict:
        return JsonResponse({"status": False, "error": "数据不存在"})
    data_dict = {
        "status": True,
        "data": obj_dict
    }
    return JsonResponse(data_dict)


def delete(request):
    """ 删除数据 """
    uid = request.GET.get("uid")
    if not models.SalesIndicator.objects.filter(id=uid).exists():
        return JsonResponse({"status": False, "error": "删除失败，数据不存在"})
    models.SalesIndicator.objects.filter(id=uid).delete()
    return JsonResponse({"status": True})


def delete_all(request):
    """ 删除所有数据 """
    pwd = request.POST.get("pwd", "")
    if not models.Admin.objects.filter(username="admin").exists():
        return JsonResponse({"status": False, "error": "超级管理员不存在"})
    password = models.Admin.objects.filter(username="admin").values().first().get("password")
    if md5(pwd) == password:
        models.SalesIndicator.objects.all().delete()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": "密码错误！"})


def addform(request):
    """ 批量上传（excel文件） """
    from openpyxl import load_workbook
    # 1.获取用户上传的文件对象
    file_obj = request.FILES.get("upload_file")
    if not file_obj:
        return JsonResponse({"status": False, "error": "请上传文件"})

    # 2.对象传递给openpyxl，又openpyxl读取文件的内容
    try:
        wb = load_workbook(file_obj)
        sheet = wb.active  # 取第一个工作表
    except Exception:
        return JsonResponse({"status": False, "error": "Excel 解析失败"})

    # 3.循环获取每一行数据
    sales_data_list = []
    error_list = []
    for idx, row in enumerate(sheet.iter_rows(min_row=3), start=2):  # 从第3行开始，遍历索引从2开始
        salesperson, created = models.Salesperson.objects.get_or_create(name=row[0].value)  # 支持自动创建不存在的 Salesperson
        date_dict = {
            'name': salesperson,  # 赋值实例
            'year': row[1].value,
            'target_sales_volume_1': row[2].value,
            'target_sales_volume_2': row[3].value,
            'target_sales_volume_3': row[4].value,
            'target_sales_volume_4': row[5].value,
            'target_sales_volume_5': row[6].value,
            'target_sales_volume_6': row[7].value,
            'target_sales_volume_7': row[8].value,
            'target_sales_volume_8': row[9].value,
            'target_sales_volume_9': row[10].value,
            'target_sales_volume_10': row[11].value,
            'target_sales_volume_11': row[12].value,
            'target_sales_volume_12': row[13].value,
            'target_sales_revenue_1': row[14].value,
            'target_sales_revenue_2': row[15].value,
            'target_sales_revenue_3': row[16].value,
            'target_sales_revenue_4': row[17].value,
            'target_sales_revenue_5': row[18].value,
            'target_sales_revenue_6': row[19].value,
            'target_sales_revenue_7': row[20].value,
            'target_sales_revenue_8': row[21].value,
            'target_sales_revenue_9': row[22].value,
            'target_sales_revenue_10': row[23].value,
            'target_sales_revenue_11': row[24].value,
            'target_sales_revenue_12': row[25].value,
        }

        # 必填字段校验
        try:
            form = SalesIndicatorModelForm(date_dict)  # 用 ModelForm 进行校验
            if form.is_valid():
                sales_data_list.append(models.SalesIndicator(**date_dict))
            else:
                error_list.append(f"第 {idx} 行错误: {form.errors}")
        except Exception as error:
            error_list.append(f"第 {idx} 行错误: <ul><li>{error}(查看日期格式是否规范化，标准2025-01-20)</li></ul>")

    # 如果有错误，返回错误信息
    if error_list:
        return JsonResponse({"status": False, "error": error_list})

    try:
        models.SalesIndicator.objects.bulk_create(sales_data_list)
    except Exception as error:
        print(error)
        return JsonResponse({"status": False, "error": f"数据导入数据库失败: {str(error)}"})
    return JsonResponse({"status": True})
