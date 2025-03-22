import time
from django.http import JsonResponse
from django.shortcuts import render, redirect, HttpResponse
from django.utils.safestring import mark_safe

from app01 import models
from app01.utils.form import SalesDataModelForm
from app01.utils.form_btn_add import FormBtnAdd
from app01.utils.form_btn_edit import FormBtnEdit
from app01.utils.form_btn_delete import FormBtnDelete
from app01.utils.form_btn_upload import FormBtnUpload
from app01.utils.form_btn_search import FormBtnSearch
from app01.utils.pagination import Pagination


def salesdata_list(request):
    """ 销售大信息列表 """
    form = SalesDataModelForm()  # 引入表模板
    search_bar = FormBtnSearch(request, form)  # 搜索框组件对象
    queryset = models.SalesData.objects.filter(**search_bar.filter()).order_by("-id")
    page_obj = Pagination(request, queryset, "page", 15)  # 翻页组件对象
    add_obj = FormBtnAdd(request, form, "/salesdata/add", "新增销售数据")  # 添加框组件对象
    edit_obj = FormBtnEdit(request, form, "/salesdata/edit_detail", "/salesdata/edit", "编辑销售数据")  # 编辑框组件对象
    del_obj = FormBtnDelete("/salesdata/delete")  # 删除框组件对象
    upload_obj = FormBtnUpload(request, "/salesdata/addform", "导入销售信息", "选择规范化的excel文件进行导入")  # excel批量上传框组件对象
    context = {
        # 搜索框组件
        "search_Modal": search_bar.html_modal(),
        "search_js": search_bar.js(),
        # 新增数据框组件
        "add_Modal": add_obj.html_modal(),
        "add_js": add_obj.js(),
        # 编辑数据框组件
        "edit_Modal": edit_obj.html_modal(),
        "edit_js": edit_obj.js(),
        # 删除框组件
        "delete_Modal": del_obj.html_modal(),
        "delete_js": del_obj.js(),
        # excel批量上传框组件
        "upload_Modal": upload_obj.html_modal(),
        "upload_js": upload_obj.js(),
        # 翻页组件
        "queryset": page_obj.page_queryset,  # 分完页的数据
        "page_string": page_obj.html(),  # html页码
    }
    return render(request, "salesdata_list.html", context)


def salesdata_add(request):
    """ 新增销售数据 """
    form = SalesDataModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})  # 直接传入dict，django会自动转json
    return JsonResponse({"status": False, "error": form.errors})


def salesdata_edit(request):
    """ 编辑信息 """
    uid = request.GET.get("uid")
    obj = models.SalesData.objects.filter(id=uid).first()
    if not obj:
        return JsonResponse({"status": False, "tips": "无法编辑，该行数据不存在，请刷新重试"})
    form = SalesDataModelForm(data=request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


def salesdata_edit_detail(request):
    """ 获取要编辑的销售数据 """
    uid = request.GET.get("uid")
    obj_dict = models.SalesData.objects.filter(id=uid).values().first()  # 获得的是一个字典
    # obj_dict = {"tittle": "1", "price": "22", "status": "1"}
    if not obj_dict:
        return JsonResponse({"status": False, "error": "数据不存在"})
    data_dict = {
        "status": True,
        "data": obj_dict
    }
    return JsonResponse(data_dict)


def salesdata_delete(request):
    """ 删除销售数据 """
    uid = request.GET.get("uid")
    if not models.SalesData.objects.filter(id=uid).exists():
        return JsonResponse({"status": False, "error": "删除失败，数据不存在"})
    models.SalesData.objects.filter(id=uid).delete()
    return JsonResponse({"status": True})


def salesdata_addform(request):
    """ 批量上传（excel文件） """
    from openpyxl import load_workbook
    # 1.获取用户上传的文件对象
    file_obj = request.FILES.get("upload_file")
    if not file_obj:
        return JsonResponse({"status": False, "error": "请上传文件"})

    # 2.对象传递给openpyxl，又openpyxl读取文件的内容
    try:
        wb = load_workbook(file_obj)
        sheet = wb.active  # 取第一个工作表
    except Exception:
        return JsonResponse({"status": False, "error": "Excel 解析失败"})

    # 3.循环获取每一行数据
    sales_data_list = []
    error_list = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        date_dict = {
            'date': row[0].value,
            'order_number': row[1].value,
            'client_company': row[2].value,
            'k3': row[4].value,
            'product_name': row[5].value,
            'product_specification': row[6].value,
            'sales_volume': row[8].value,
            'department': row[10].value,
            'salesperson': row[11].value,
            'gross_unit_price': row[12].value,
            # 销售金额（元）= 实发数量 * 销售单价(元/Kg）
            'net_unit_price': row[14].value,
            # 未税金额（元）= 实发数量 * 不含税单价（元/Kg）
            'actual_client_company': row[16].value,
            'supply_company': row[18].value,
            'intra_group_or_external_sales': row[19].value,
            'product_domain_groups': row[21].value,
            'business_trade_categories': row[22].value,
            'new_and_returning_customers': row[23].value,
            'product_category': row[30].value,
            'core_product': row[31].value,
            'order_type': row[32].value,
        }
        # 必填字段校验
        try:
            form = SalesDataModelForm(date_dict)  # 用 ModelForm 进行校验
            if form.is_valid():
                sales_data_list.append(models.SalesData(**date_dict))
            else:
                error_list.append(f"第 {idx} 行错误: {form.errors}")
        except Exception as error:
            error_list.append(f"第 {idx} 行错误: <ul><li>{error}(查看日期格式是否规范化，标准2025-01-20)</li></ul>")

    # 如果有错误，返回错误信息
    if error_list:
        return JsonResponse({"status": False, "error": error_list})

    try:
        models.SalesData.objects.bulk_create(sales_data_list)
    except Exception as error:
        print(error)
        return JsonResponse({"status": False, "error": f"数据导入数据库失败: {str(error)}"})
    return JsonResponse({"status": True})
